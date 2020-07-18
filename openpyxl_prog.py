import openpyxl as op

drr=op.load_workbook('D:/DSR/DRR.xlsx')
iurr=op.load_workbook('D:/DSR/IURR.xlsx')

drrws=drr.active
iurrws=iurr.active
final=op.Workbook()
final.copy_worksheet()

final.save(filename='D:/DSR/final.xlsx')
